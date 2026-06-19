#####################################################################
#   稼働データ 月次推移表(ViewTable)作成Module Ver.1.0
#   Last Update on 2023.5.4
#####################################################################
# インポート処理
import openpyxl
import os
from datetime import datetime,timedelta
import configparser

# 定数定義
OUTPUT_FOLDER = '\\\\LS720D6C6\\share\\P_ProductControl\\OperationData\\'   #出力フォルダ

#グローバル変数
wb = None   # Excelブック格納用
view_sheets = ['Fact1','Fact2_3','Fact4']

#----- 関数定義 ---------------------------------------------------------
def make_table(yearmonth:str,workbook)->None:
    """ ViewTable作成メインプロセス"""
    global wb
    wb = workbook   # データ用エクセルファイル格納
    
    sheet_names = wb.sheetnames #全シート名取得
    for sht in view_sheets:
        if sht not in sheet_names:
            print(f'--- シート[{sht}]が見つかりません。 ---')
            print('Viewシートは次回自動作成されます。')
            return
        else:
            input_date(sht,yearmonth)   #Viewシートへ日付入力
            input_to_vsheet(sht)        #Viewシートへ稼働データ入力
            move_sheet_to_end(sht)      #シートを末尾へ移動


def input_date(sheet_name:str, yearmonth:str)->None:
    """ ViewTableシートへの日付入力処理
        Args:
            sheet_name(str) : ViewTableシート名称
            yearmonth(str) : Excelファイルの該当年月(西暦下2桁+月2桁 )
    """
    date_from = datetime.strptime(yearmonth+'01','%y%m%d')
    for i in range(31):
        input_date = date_from + timedelta(days=i)
        wb[sheet_name].cell(3,4+i).value = input_date   #日産数用
        wb[sheet_name].cell(21,4+i).value = input_date  #稼働率用
        wb[sheet_name].cell(39,4+i).value = input_date  #異常回数用


def input_to_vsheet(sheet_name:str)->None:
    """指定されたViewTableに自動入力する"""
    print('--- Start processing ---')
    view_sht = wb[sheet_name]   #View用シート選択
    for row in range(1,100):    # row : 機械番号検索用(列方向)
        val = str(view_sht.cell(row,1).value)
        if '号機' in val:   # 機械番号入力時
            mcno = val.replace('号機','')
            dclass = view_sht.cell(row,3).value
        else:
            continue

        for col in range(4,35): # col : 日付データ検索用(行方向)
            val = view_sht.cell(3, col).value
            if val != None: #日付データ入力されている場合
                input_date = val.strftime('%y%m%d') #入力日付取得

                # データ入力セルの格納
                print(f'mcno={mcno}/input_date={input_date}/dclass={dclass}'
                    ,f'/target_data={get_ope_data(mcno,input_date,dclass)}')
                ope_data = get_ope_data(mcno,input_date,dclass)
                view_sht.cell(row,col).value = ope_data
                # セル背景塗りつぶし処理
                fill = openpyxl.styles.PatternFill(
                    patternType='solid',
                    fgColor=get_fill_color(ope_data,mcno,input_date,dclass))
                view_sht.cell(row,col).fill = fill


def get_fill_color(opdata,mcno,date,dclass)->str:
    """稼働率データの値により背景塗りつぶし色を返却する"""
    # 定数(色の16進数)←Excelカラーパレットで調査
    WHITE='FFFFFF'; PINK = 'FF99FF'; YELLOW = 'FFFF66'
    GREEN = '99FF99'; BLUE = '99CCFF'; GRAY = 'D9D9D9'

    # 土日はGRAY
    jg_date = datetime.strptime(date,'%y%m%d')
    weekday = jg_date.weekday()
    if weekday==5 or weekday==6: #土日の場合
        return GRAY

    # データがない場合(空文字/None)は白
    if opdata=='' or opdata==None:
        return WHITE

    # 日産数データ(Production Numbers)
    if dclass=='PN':
        goal = get_prod_goal(mcno)
        print(f'McNo={mcno}/ProdGoal={goal}')
        if opdata >= goal:
            return BLUE
        elif opdata >= int(goal*0.75):
            return GREEN
        elif opdata >= int(goal*0.5):
            return YELLOW
        else:
            return PINK
        
    # 稼働率データ(Operation Rate)
    if dclass=='OR':
        if opdata >= 85:
            return BLUE
        elif opdata >= 60:
            return GREEN
        elif opdata >= 40:
            return YELLOW
        else:
            return PINK
    
    # アラーム発生回数(Alarm Numbers)
    if dclass=='AN':
        if opdata <= 3:
            return BLUE
        elif opdata <= 6:
            return GREEN
        elif opdata <= 10:
            return YELLOW
        else:
            return PINK

    return WHITE    # 上記以外の場合


def get_prod_goal(mcno:str) -> int:
    """INIファイルから機械の目標生産数を取得する"""
    inifile = configparser.ConfigParser()
    inifile.read('config.ini','UTF-8')
    for blk_no in range(80):     # 2025/6/2変更 50→80
        mcno_ini = inifile.get('block'+str(blk_no).zfill(3),'mc_no').replace('No.','')
        if mcno == mcno_ini:
            return int(inifile.get('block'+str(blk_no).zfill(3),'prod_goal'))
    print('---ERROR : Production goal is not found---')
    return 0


def get_ope_data(mcno:str, date:str, dclass:str)->str:
    """機械番号・日付(シート名)・データ種類から目的のデータを取得"""
    # シートがあるか確認→ないならから空文字列
    sheet_names = wb.sheetnames #全シート名取得
    if date not in sheet_names:
        # print(f'<sheet={date} is not found>') # ForDebug
        return ''
    
    sht = wb[date]  # シート選択
    
    #シート中の機械番号検索(2行目)＆目的データ取得
    data_row = 0    # 各種稼働データ格納行
    for col in range(2,100):    # 100列目までサーチ
        val = str(sht.cell(3,col).value).replace('No.','')    #機械番号の数字のみ
        if mcno == val: # 対象機械の列の場合
            if dclass=='PN':    # 日産数
                data_row = 6    
            elif dclass=='OR':  # 定時稼働率
                data_row = 17    #<要修正> 現状データなし
            elif dclass=='AN':  # 異常発生回数
                data_row = 7
            else:
                print('<Data-class not input>')
                return ''
            # 該当データの返却
            return sht.cell(data_row,col).value


def move_sheet_to_end(sheet_name:str)->None:
    """エクセルシートを末尾へ移動する"""
    # <注> openpyxlは相対的な移動(offset)しかできない
    # 本処理実行後、複数のシートが選択される現象が発生する(が、無視する)
    sht = wb[sheet_name]
    sht_num = 0
    sht_pos = 0 # 対象シート位置
    for s in wb.worksheets:
        sht_num += 1
        if sht==s:
            sht_pos = sht_num   #対象シート位置格納
        else:
            continue
    wb.move_sheet(sht, offset=sht_num-sht_pos)  # シート移動


# モジュール単体テスト用コード ------------------------------------------
# <注意>
#  通常は本プログラムはPLCと通信したタイミングで自動実行される。
#  過去のデータに対して処理したいときは下記のyearmonthを変更して実行する。
if __name__=='__main__':

    yearmonth = '2304'  # 対象年月設定(YYmm形式)

    # カレントディレクトリ変更(VSCode使用時)
    os.chdir(os.path.dirname(__file__))


    fname = 'OpeData-' + yearmonth + '.xlsx'
    fpath = OUTPUT_FOLDER + fname

    # データファイルオープン処理
    if not os.path.isfile(fpath):
        print(f'データファイル{fname}が見つかりません')
        exit()
    else:
        print('--- Loading excel file ---')
        wb = openpyxl.load_workbook(fpath)
        # <注> data_only=Trueで数式の演算結果を取得できる
    
    # メインプロセス
    make_table(yearmonth,wb)

    print('--- Saving work book ---')
    wb.save(fpath)



#### MEMO #################################################
# Openpyxlはブック間のシートのコピーができない
# →別ブックからの値のコピーはできるが、書式のコピーができない。
# →xlwingsを使った方が良さそう
# 
# Openpyxlでload_workbook実行時にdata_only=Trueとすると、
# 数式ではなく演算結果が取得できる。
# 但し、save実行時に全ての数式が値に変換される→使えん。。。。
#
#
#
#
#
