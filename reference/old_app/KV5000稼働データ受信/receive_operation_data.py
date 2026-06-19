#####################################################################
#   KV-5000 稼働データ受信プログラム Ver.1.0
#   Last Update on 2023.3.26
#####################################################################
# インポート処理
import socket
import openpyxl
import os
import configparser
import subprocess
from datetime import datetime
import make_view_table      # 独自モジュール
import copy_view_table      # 独自モジュール

# 関数定義 --------------------------------------------------------------
# PLC通信処理
def com_with_plc(ip_add,port_no,cmd):
    server = (ip_add,port_no)
    skt = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    skt.connect(server)

    # コマンド送信
    skt.send(cmd.encode('ASCII'))
    # レスポンス受信8192
    res = skt.recv(8192).decode()
    skt.close()
    return res


# メインプロセス ---------------------------------------------------------------------

os.chdir(os.path.dirname(__file__))

# 定数定義
#IP_ADDRESS = '192.168.1.111'    # PLCのIPアドレス 
IP_ADDRESS = '172.20.1.111'    # PLCのIPアドレス
PORT_NO = 8501  # ポート番号(PLC設定に合わせる)
OUTPUT_FOLDER = '//LS720D6C6/share/P_ProductControl/OperationData/'   #出力フォルダ


# グローバル変数定義
date = ''       # 対象データの日付(YYmmdd形式)
out_fname = ''  # データ出力用Excelファイル名(OpeData-YYmm.xlsx)
out_fpath = ''  # データ出力用Excelファイル フルパス

# データ出力用ファイル&シートの準備
if True:
    # データの日付受信(ZF0の32bitでYYmmdd形式で格納)
    cmd = 'RD ZF0.D\r'
    res = com_with_plc(IP_ADDRESS,PORT_NO,cmd)
    print(res)
    date = res[4:10]    #YYmmdd形式
    print(f'対象データ日付 = {date}') # ForDebug

    # データ格納用エクセルファイル名設定
    out_fname = 'OpeData-' + date[:4] + '.xlsx'
    print(f'出力用ファイル = {out_fname}')

    # ファイルパスの格納
    out_fpath = OUTPUT_FOLDER + out_fname

    # ViewTable準備(template.xlsxからシートコピー)
    # (シートのコピーはopenpyxlは不可→xlwings使う)
    # (データファイルがない場合は、スキップ)
    if os.path.isfile(out_fpath):
        copy_view_table.copy_view_table(out_fpath)

    # データ格納用ファイル有無確認＆新規作成
    if os.path.isfile(out_fpath):
        print('--- output-file exists ---')
        # 入力用シート有無確認&新規作成
        wb = openpyxl.load_workbook(out_fpath)
        if date in wb.sheetnames:
            print('--- output-sheet exists ---')
        else:
            print('--- Creating new sheet ---')
            wb.create_sheet(date)
            wb.save(out_fpath)
    else:
        print('--- Creating new output-file ---')
        wb = openpyxl.Workbook()
        wb.create_sheet(date)
        wb.remove(wb['Sheet'])  # デフォルトシートの削除
        wb.save(out_fpath)

# データ出力用WorkbookとSheet設定
wb = openpyxl.load_workbook(out_fpath)
ws = wb[date]

# フォーマット準備
if True:
    print('--- Input title data ---')
    # 日付入力
    ws.cell(1,1).value = datetime.now().strftime('%Y/%m/%d %H:%M')
    # 稼働データ項目の入力(A列)
    op_data_title = ['稼働データ項目','機械番号','機械型式','日産数-目標',
                     '日産数-実績','異常発生数','全稼働時間/分',
                     '定時稼働時間/分','単純停止時間/分','異常停止時間/分',
                     '刃具交換時間/分','段替え時間/分','故障停止時間/分',
                     '材料切れ時間/分','不明時間/分','定時稼働率/%']
    for n,x in enumerate(op_data_title):
        ws.cell(n+2,1).value = x    # 2行目から入力

    # ZFデバイス番号入力(A列)
    for i in range(1500):
        ws.cell(i+30,1).value = 'ZF' + str(i).zfill(4)

    # BlockNo入力(1行)
    for i in range(80):     # 2025/6/2変更 50→80
        ws.cell(2,i+2).value = 'Blk.' + str(i)


# config.iniからの転記(機械No/機械型式/目標日産数)
if True:
    print('--- input from config.ini ---')
    inifile = configparser.ConfigParser()
    inifile.read('config.ini','UTF-8')
    for blk_no in range(80):     # 2025/6/2変更 50→80
        # 機械No入力
        val = inifile.get('block'+ str(blk_no).zfill(3),'mc_no')
        ws.cell(3,blk_no+2).value = val
        # 機械型式入力
        val = inifile.get('block'+ str(blk_no).zfill(3),'mc_type')
        ws.cell(4,blk_no+2).value = val
         # 目標日産数入力
        val = inifile.get('block'+ str(blk_no).zfill(3),'prod_goal')
        ws.cell(5,blk_no+2).value = int(val)
       
# 生データ受信&稼動データ入力(50ブロック分)
if True:
    for blk_no in range(80):     # 2025/6/2変更 50→80
        print(f'--- BlockNo={blk_no}:Process Start ---')
        # 1Block(1500ワード受信処理)
        if True:
            data=[] # 受信データ格納用
            # 通信1回目(前1000データ)
            cmd = 'RDS ZF' + str(blk_no*1500) + '.U 1000\r'
            res = com_with_plc(IP_ADDRESS,PORT_NO,cmd)
            data = res.split()
            # 通信2回目(後ろ500データ)
            cmd = 'RDS ZF'+ str(blk_no*1500+1000) +'.U 500\r'
            res = com_with_plc(IP_ADDRESS,PORT_NO,cmd)
            data += res.split()
        
        # 生データ格納(30行目から)
        if True:
            for n,x in enumerate(data):
                ws.cell(n+30,blk_no+2).value = int(x)

        # 各種稼働データ算出&入力
        # (リストの内包表記とスライスに注意すること)
        if True:
            # 日産数-実績(ZF2)
            ws.cell(6,blk_no+2).value = int(data[2])
            # 異常発生数(ZF3)
            ws.cell(7,blk_no+2).value = int(data[3])
            # 全稼働時間(4:00=ZF10 - 3:59=ZF1449 StatusNo=15)
            val = len([x for x in data[10:1450] if int(x)==15])
            ws.cell(8,blk_no+2).value = val
            # 定時稼働時間(8:00=ZF250 - 16:59=ZF789 StatusNo=15)
            val = len([x for x in data[250:790] if int(x)==15])
            ws.cell(9,blk_no+2).value = val
            # 単純停止時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=16)
            val = len([x for x in data[250:790] if int(x)==16])
            ws.cell(10,blk_no+2).value = val
            # 異常停止時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=20)
            val = len([x for x in data[250:790] if int(x)==20])
            ws.cell(11,blk_no+2).value = val
            # 刃具交換時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=1)
            val = len([x for x in data[250:790] if int(x)==1])
            ws.cell(12,blk_no+2).value = val
            # 段替え時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=2)
            val = len([x for x in data[250:790] if int(x)==2])
            ws.cell(13,blk_no+2).value = val
            # 故障停止時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=3)
            val = len([x for x in data[250:790] if int(x)==3])
            ws.cell(14,blk_no+2).value = val
            # 材料切れ時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=4)
            val = len([x for x in data[250:790] if int(x)==4])
            ws.cell(15,blk_no+2).value = val
            # 不明時間(定時内)(8:00=ZF250 - 16:59=ZF789 StatusNo=0)
            val = len([x for x in data[250:790] if int(x)==0])
            ws.cell(16,blk_no+2).value = val
            # 定時稼働率(小数点以下1桁)
            val = round((ws.cell(9,blk_no+2).value /540*100),1)
            ws.cell(17,blk_no+2).value = val

# Excelファイル書式設定等
if True:
    print('--- Adjusting Excel Layout ---')
    # Excelシート ウィンドウ枠の固定(A列が固定される)
    ws.freeze_panes = 'B1'
    # 列幅調整
    ws.column_dimensions['A'].width = 16
    for i in range(2,52):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 7.5

# 稼働データ一覧表(ViewTable)への入力
make_view_table.make_table(date[:4],wb)


print('--- Saving Workbook ---')
wb.save(out_fpath)



# 確認のためExcelファイルオープン
res = input('エクセルファイル開きますか？(yes=y)')
if res=='y':
    # コマンドプロンプトは「バックスラッシュ(\)」にする必要あり
    out_fpath = out_fpath.replace('/','\\')
    subprocess.Popen(['start',out_fpath],shell=True)


### MEMO ############################################################
#
# 《残作業》
#　・他の稼動データ計算(稼働率/可動率など)
#　・エクセルシート並べ替え
#
#　・機械ごとの推移
#　・主要データグラフ化
#
#　・項目が増えた場合に行位置を自動で対応するようにする
#　（もしくはconfig.ini管理）
#